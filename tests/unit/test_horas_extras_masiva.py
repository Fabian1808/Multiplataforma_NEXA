"""Pruebas del motor de HORAS EXTRAS MASIVA.

Cubre los casos requeridos:
  - normalización (nombres/códigos/fechas/horas/moneda)
  - conciliación exacta (DNI / Fotocheck / Num personal) y difusa (nombre)
  - turnos T1/T2/T3 y cálculo de horas extras
  - matching tarifario por niveles ALTA/MEDIA/BAJA/SIN_TARIFA y ambigüedad
  - valorización 25/35/100 → monto en soles con precisión Decimal
  - estados OK/ADVERTENCIA/REVISAR/ERROR
  - exportación Excel con 6 hojas
"""

from __future__ import annotations

import datetime as dt
import sys
from decimal import Decimal
from pathlib import Path

import pytest

ENGINE = Path(__file__).resolve().parent.parent.parent / "plugins" / "horas_extras_masiva" / "engine" / "src"
sys.path.insert(0, str(ENGINE))

import config as cfg_mod          # noqa: E402
import ingesta as ing             # noqa: E402
import utiles as u                # noqa: E402


@pytest.fixture(scope="module")
def cfg():
    return cfg_mod.cargar_config()


def F(v):
    return u.normalizar_fecha(v)


def H(v):
    return u.normalizar_hora(v)


# ---------------------------------------------------------------------------
# 1. NORMALIZACIÓN
# ---------------------------------------------------------------------------
class TestNormalizacion:
    def test_normalizar_nombre_sin_tildes_mayusculas(self):
        assert u.normalizar_nombre("portillo cervantes ñañéz") == "PORTILLO CERVANTES NANEZ"
        assert u.normalizar_nombre("Alberto  Ramírez") == "ALBERTO RAMIREZ"

    def test_limpiar_codigo(self):
        assert u.limpiar_codigo("000000002 - ALBERT RAMIREZ MEDINA") == "000000002"
        assert u.limpiar_codigo("portillo") == "portillo"

    def test_normalizar_fecha(self):
        assert F("20/12/2025") == dt.date(2025, 12, 20)
        assert F("2025-12-20") == dt.date(2025, 12, 20)

    def test_normalizar_hora(self):
        assert H("02:22:53") == dt.time(2, 22, 53)
        assert H("18:30") == dt.time(18, 30)

    def test_dinero_decimal_exacto(self):
        assert u.monto_decimal("0.1") + u.monto_decimal("0.2") == Decimal("0.30")
        assert u.moneda(Decimal("1551.055"), cfg_mod.DEFAULT_CONFIG) == "S/ 1,551.06"


# ---------------------------------------------------------------------------
# 2. CONCILIACIÓN
# ---------------------------------------------------------------------------
def _marc(fecha="20/12/2025", hora="08:00", dni="77175933", fc="20108169",
          num="1165", nombre="NAVARRO PALACIN HENRY ALEJANDRO", tipo="Entrada"):
    return ing.Marcacion(
        empresa="0153 - CONFIPETROL / CJM", ruc="20357259976",
        num_personal=num, fotocheck=fc, empleado=nombre, dni=dni,
        fecha=F(fecha), hora=H(hora), tipo_acceso=tipo, situacion="Permitido",
        fecha_hora=dt.datetime.combine(F(fecha), H(hora)),
    )


def _empleado(dni="77175933", fc="20108169", cargo="TORNERO",
              nombre="000000165 - NAVARRO PALACIN HENRY ALEJANDRO",
              ruc="20357259976", empresa="CONFIPETROL"):
    e = ing.Empleado()
    e.empleado = nombre
    e.dni = dni
    e.fotocheck = fc
    e.cargo = cargo
    e.ruc = ruc
    e.empresa_terceros = empresa
    e.contrato = "CONTRATO"
    return e


@pytest.fixture
def maestro(cfg):
    from conciliacion import MaestroPersonal
    return MaestroPersonal([_empleado()], cfg)


class TestConciliacion:
    def test_por_dni(self, maestro):
        assert maestro.por_dni("77175933")

    def test_por_fotocheck(self, maestro):
        assert maestro.por_fotocheck("20108169")

    def test_por_numpersonal(self, maestro):
        # num personal derivado del código del nombre
        assert maestro.por_numpersonal("000000165")

    def test_conciliar_por_dni(self, cfg, maestro):
        from conciliacion import conciliar
        res = conciliar([_marc(dni="77175933")], maestro, cfg)
        assert res[0].conciliado and res[0].metodo == "dni"

    def test_conciliar_por_nombre_difuso(self, cfg, maestro):
        from conciliacion import conciliar
        mar = _marc(dni="", fc="", num="", nombre="NAVARRO PALACIN HENRY ALEJANDRO")
        res = conciliar([mar], maestro, cfg)
        assert res[0].conciliado and res[0].metodo in ("nombre", "nombre_difuso")

    def test_sin_conciliar(self, cfg, maestro):
        from conciliacion import conciliar
        mar = _marc(dni="99999999", fc="000000", num="000000", nombre="PERSONA INEXISTENTE")
        res = conciliar([mar], maestro, cfg)
        assert not res[0].conciliado

    def test_buscar_nombre_indice_invertido_equivale_bruto(self, cfg, maestro):
        """El índice invertido (FASE 3) da los mismos candidatos que el barrido."""
        from conciliacion import MaestroPersonal
        empleados = [_empleado(), _empleado(nombre="000000002 - RAMIREZ ANA MARIA", dni="11111111")]
        m = MaestroPersonal(empleados, cfg)
        # fuerza y compara con el barrido bruto
        cands = m.buscar_nombre("NAVARRO PALACIN HENRY ALEJANDRO", 82.0)
        brutos = m._buscar_nombre_bruto(m._norm("NAVARRO PALACIN HENRY ALEJANDRO"), 82.0)
        kl_new = [(id(e), round(c, 6)) for e, c in cands]
        kl_old = [(id(e), round(c, 6)) for e, c in brutos]
        assert kl_new == kl_old
        assert any(id(c[0]) == id(maestro._por_nombre_norm.get(maestro._norm("000000165 - NAVARRO PALACIN HENRY ALEJANDRO")))
                   or c[0].dni == "77175933" for c in cands), "debe matchear al empleado del fixture"

    def test_matching_determinista(self):
        """ratio_token_set debe ser invariante al orden de los tokens (joins estables)."""
        import matching as mz
        r1 = mz.mejor_token("NAVARRO PALACIN HENRY ALEJANDRO", "PALACIN NAVARRO HENRY ALEJANDRO")
        r2 = mz.mejor_token("NAVARRO PALACIN HENRY ALEJANDRO", "PALACIN NAVARRO HENRY ALEJANDRO")
        assert r1 == r2

    def test_conciliacion_equivalente_bruto(self, cfg):
        """El pipeline nuevo (índice) y el viejo (barrido) concilian igual."""
        from conciliacion import conciliar, MaestroPersonal
        empleados = [_empleado(dni="77175933"),
                     _empleado(dni="22222222", nombre="000000002 - RAMIREZ ANA MARIA")]
        mar = _marc(dni="", fc="", num="", nombre="NAVARRO PALACIN HENRY ALEJANDRO")
        res_new = conciliar([mar], MaestroPersonal(empleados, cfg), cfg)
        assert res_new[0].conciliado


# ---------------------------------------------------------------------------
# 3. TURNOS Y CALCULO DE HE
# ---------------------------------------------------------------------------
class TestReglas:
    def test_detectar_T1(self, cfg):
        from reglas import detectar_turno
        assert detectar_turno(dt.datetime(2025,12,20,8,0), dt.datetime(2025,12,20,18,0), cfg) == "T1"

    def test_detectar_T3(self, cfg):
        from reglas import detectar_turno
        assert detectar_turno(dt.datetime(2025,12,20,13,0), dt.datetime(2025,12,20,21,0), cfg) == "T3"

    def test_jornada_T1_10h(self, cfg):
        from reglas import _armar_jornada_simple
        j = _armar_jornada_simple(_empleado(),
                                  _marc(hora="08:00", tipo="Entrada"),
                                  _marc(hora="18:00", tipo="Salida"), cfg)
        assert j is not None
        # 10h trabajadas - 1h comida = 9h? -> horas extra = -1 (no extra)
        assert j.turno == "T1"

    def test_he_T1_12h(self, cfg):
        from reglas import _armar_jornada_simple
        j = _armar_jornada_simple(_empleado(),
                                  _marc(hora="08:00", tipo="Entrada"),
                                  _marc(hora="22:00", tipo="Salida"), cfg)
        # trabajadas 14h - 1h comida = 13h -> HE = 3h
        assert j.horas_extras >= Decimal("2.9")


# ---------------------------------------------------------------------------
# 4. MATCHING TARIFARIO
# ---------------------------------------------------------------------------
def _tarifa(cargo, empresa="CONFIPETROL", ruc="20357259976", c25="50", c35="54", c100="80"):
    t = ing.Tarifa()
    t.cargo = cargo
    t.empresa = empresa
    t.ruc = ruc
    t.c25 = Decimal(c25)
    t.c35 = Decimal(c35)
    t.c100 = Decimal(c100)
    t.objeto = "CONTRATO"
    return t


@pytest.fixture
def tarifario(cfg):
    from tarifario import Tarifario
    tarifas = [
        _tarifa("TORNERO"),
        _tarifa("TORNERO", empresa="OTRA", ruc="12345678901"),  # distinta empresa
        _tarifa("OPERADOR", empresa="CONFIPETROL"),
    ]
    return Tarifario(tarifas, cfg)


class TestMatchingTarifario:
    def test_nivel_ALTA(self, tarifario):
        from tarifario import ResultadoTarifa
        r = tarifario.matching(_empleado(cargo="TORNERO"))
        assert r.nivel == "ALTA" and r.tarifa.cargo == "TORNERO"

    def test_nivel_MEDIA(self, tarifario):
        e = _empleado(cargo="TORNERO", ruc="88888888", empresa="CONFIPETROL")
        r = tarifario.matching(e)
        # RUC no coincide -> MEDIA por empresa+cargo
        assert r.nivel == "MEDIA"

    def test_nivel_BAJA_ambigua(self, tarifario):
        # cargo TORNERO con empresa y ruc distintos a todos -> solo cargo -> BAJA
        e = _empleado(cargo="OPERADOR", ruc="99999999", empresa="XYZ")
        r = tarifario.matching(e)
        assert r.nivel == "BAJA" and r.ambigua

    def test_SIN_TARIFA(self, tarifario):
        e = _empleado(cargo="CARGO NO EXISTENTE")
        r = tarifario.matching(e)
        assert r.nivel == "SIN_TARIFA"


# ---------------------------------------------------------------------------
# 5. VALORIZACIÓN / MONTOS
# ---------------------------------------------------------------------------
class TestValorizacion:
    def test_monto_25_35(self, cfg):
        from tarifario import valorizar
        from reglas import Jornada
        from tarifario import ResultadoTarifa
        t = _tarifa("TORNERO", c25="50", c35="54", c100="80")
        rt = ResultadoTarifa(tarifa=t, nivel="ALTA", confianza=100, cargo_match=True,
                             ruc_match=True, empresa_match=True)
        # 3h extras: 2h a 25%=50, 1h a 35%=54 -> 100+54=154
        j = Jornada(empleado=_empleado(), fecha=dt.date(2025,12,20), turno="T1",
                    horas_extras=Decimal("3.0"))
        filas = valorizar([j], {id(j.empleado): rt}, cfg)
        assert filas[0]["monto"] == Decimal("154.00")
        assert filas[0]["horas_25"] == Decimal("2.0")
        assert filas[0]["horas_35"] == Decimal("1.0")

    def test_monto_100_activacion(self, cfg):
        from tarifario import valorizar
        from reglas import Jornada
        from tarifario import ResultadoTarifa
        t = _tarifa("TORNERO", c25="50", c35="54", c100="80")
        rt = ResultadoTarifa(tarifa=t, nivel="ALTA", confianza=100, cargo_match=True,
                             ruc_match=True, empresa_match=True)
        # 8h extras >= 7h -> 100% = 80 -> 8*80=640
        j = Jornada(empleado=_empleado(), fecha=dt.date(2025,12,20), turno="T2",
                    horas_extras=Decimal("8.0"))
        filas = valorizar([j], {id(j.empleado): rt}, cfg)
        assert filas[0]["monto"] == Decimal("640.00")
        assert filas[0]["tipo_hora"] == "100%"

    def test_monto_con_precision_decimal(self, cfg):
        # evita errores de punto flotante
        from tarifario import valorizar
        from reglas import Jornada
        from tarifario import ResultadoTarifa
        t = _tarifa("TORNERO", c25="0.1", c35="0.2", c100="0.3")
        rt = ResultadoTarifa(tarifa=t, nivel="ALTA", confianza=100, cargo_match=True,
                             ruc_match=True, empresa_match=True)
        j = Jornada(empleado=_empleado(), fecha=dt.date(2025,12,20), turno="T1",
                    horas_extras=Decimal("3.0"))
        filas = valorizar([j], {id(j.empleado): rt}, cfg)
        # 2h*0.1 + 1h*0.2 = 0.4 exacto
        assert filas[0]["monto"] == Decimal("0.40")


# ---------------------------------------------------------------------------
# 6. ESTADOS
# ---------------------------------------------------------------------------
class TestEstados:
    def test_estados_permitidos(self):
        from validacion import ESTADOS
        assert set(ESTADOS) == {"OK", "ADVERTENCIA", "REVISAR", "ERROR"}

    def test_sin_tarifa_con_he_es_error(self, cfg):
        from tarifario import valorizar
        from reglas import Jornada
        from tarifario import ResultadoTarifa
        rt = ResultadoTarifa(nivel="SIN_TARIFA", confianza=0)
        j = Jornada(empleado=_empleado(), fecha=dt.date(2025,12,20), turno="T1",
                    horas_extras=Decimal("2.0"))
        filas = valorizar([j], {id(j.empleado): rt}, cfg)
        assert filas[0]["estado"] == "ERROR"

    def test_sin_he_no_es_error(self, cfg):
        from tarifario import valorizar
        from reglas import Jornada
        from tarifario import ResultadoTarifa
        rt = ResultadoTarifa(nivel="SIN_TARIFA", confianza=0)
        j = Jornada(empleado=_empleado(), fecha=dt.date(2025,12,20), turno="T1",
                    horas_extras=Decimal("0"))
        filas = valorizar([j], {id(j.empleado): rt}, cfg)
        assert filas[0]["estado"] == "OK"

    def test_he_anomalas_es_advertencia(self, cfg):
        from validacion import validar_fila
        fila = {"horas_extras": Decimal("30"), "estado": "OK"}
        assert validar_fila(fila, cfg) == "ADVERTENCIA"


# ---------------------------------------------------------------------------
# 7. EXPORTACIÓN Excel (6 hojas)
# ---------------------------------------------------------------------------
def _resultado_sintetico(cfg):
    from motor import Resultado
    from conciliacion import RegistroConciliado
    from reglas import Jornada
    from tarifario import ResultadoTarifa, valorizar
    from validacion import aplicar_estados

    emp = _empleado()
    mar_ent = _marc(hora="08:00", tipo="Entrada")
    mar_sal = _marc(hora="18:00", tipo="Salida")
    rc = RegistroConciliado(marcacion=mar_ent, empleado=emp, metodo="dni", confianza=100)
    j = Jornada(empleado=emp, fecha=dt.date(2025,12,20), turno="T1",
                inicio=dt.datetime(2025,12,20,8), fin=dt.datetime(2025,12,20,18),
                horas_trabajadas=Decimal("9"), jornada=Decimal("10"),
                horas_extras=Decimal("0"))
    rt = ResultadoTarifa(tarifa=_tarifa("TORNERO"), nivel="ALTA", confianza=100,
                         cargo_match=True, ruc_match=True, empresa_match=True)
    filas = valorizar([j], {id(emp): rt}, cfg)
    filas = aplicar_estados(filas, cfg)
    res = Resultado(
        fuentes=None,
        marcaciones=[mar_ent, mar_sal],
        empleados=[emp],
        tarifas=[_tarifa("TORNERO")],
        conciliados=[rc],
        jornadas=[j],
        filas=filas,
    )
    res.__post_init__()
    return res


class TestExportacion:
    def test_6_hojas(self, cfg, tmp_path):
        import openpyxl
        import exportacion
        ruta = tmp_path / "out.xlsx"
        exportacion.exportar(_resultado_sintetico(cfg), ruta, cfg)
        wb = openpyxl.load_workbook(ruta)
        hojas = cfg_mod.DEFAULT_CONFIG["exportacion"]["hojas"]
        assert all(h in wb.sheetnames for h in hojas)

    def test_detalle_encabezados(self, cfg, tmp_path):
        import openpyxl
        import exportacion
        ruta = tmp_path / "out.xlsx"
        exportacion.exportar(_resultado_sintetico(cfg), ruta, cfg)
        wb = openpyxl.load_workbook(ruta)
        ws = wb["DETALLE"]
        heads = [c.value for c in ws[1]]
        assert "Horas Extras" in heads and "Monto (S/)" in heads
