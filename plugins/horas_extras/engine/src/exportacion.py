"""Exportación de resultados: log de auditoría, histórico y consolidado.

- Log de auditoría (CSV) con las columnas definidas en la especificación.
- Resumen por período (JSON) para el histórico y el dashboard.
- Copia del consolidado con las celdas objetivo (AZ/BA/BE y las columnas de
  la plantilla: Hora de Inicio, Hora Fin, VALIDACION RAINBOW, TIPO DE HHEE)
  completadas. Nunca se modifica el archivo original.
"""

import csv
import json
import os
from datetime import datetime

import openpyxl

from impacto import ESTADOS_PENDIENTES, ESTADOS_SIN_SUSTENTO
from empresas import horas_a_hhmm
from reglas import estado_fila, observaciones_fila

# Estados sin evidencia suficiente para calcular horas extras de forma
# automatica: la columna muestra "REVISIÓN REQUERIDA" en lugar de inventar
# horas. Incluye OBSERVADO (hay marcas pero ninguna forma pareja
# entrada/salida utilizable).
ESTADOS_HE_REVISION = set(ESTADOS_PENDIENTES) | {
    "SIN MARCACIÓN", "SIN ENTRADA", "SIN SALIDA", "OBSERVADO"}

ETIQUETA_REVISION = "REVISIÓN REQUERIDA"

COLUMNAS_LOG = [
    "Fila Excel", "Empresa", "Empleado Consolidado", "Empleado Relatorio",
    "DNI", "Fecha",
    "Turno", "Hora Inicio", "Hora Final", "Fecha Salida", "Score Nombre",
    "Confianza", "Metodo Nombre", "Estado", "Observación", "Horas Declaradas",
    "Horas Validadas", "Horas Trabajadas", "Horas Netas (menos comida)",
    "Jornada Nominal", "Horas Extras",
    "Diferencia Horas", "Monto Declarado", "Monto Validado",
    "Monto Observado", "Ahorro Potencial", "Validacion Rainbow", "Tipo HHEE",
]


def _hora_fmt(value):
    return value.strftime("%H:%M:%S") if value else ""


def _fecha_fmt(value):
    return value.isoformat() if value else ""


def _monto_validado(resultado):
    if resultado.estado == "VALIDADO":
        return round(resultado.monto_total, 2)
    if resultado.estado in ESTADOS_SIN_SUSTENTO:
        return 0.0
    return "PENDIENTE"


def _horas_extras_fmt(resultado):
    """'HH:MM' de horas extras pagables, o 'REVISIÓN REQUERIDA'.

    La etiqueta aparece cuando no hay evidencia suficiente (pareja
    entrada/salida incompleta o estado inconcluso); nunca se inventan horas.
    """
    if resultado.estado in ESTADOS_HE_REVISION:
        return ETIQUETA_REVISION
    if resultado.horas_extras is not None:
        return horas_a_hhmm(resultado.horas_extras)
    return ETIQUETA_REVISION


def _fila_log(resultado):
    horas_declaradas = resultado.horas_declaradas
    horas_validadas = resultado.horas_validadas
    try:
        diferencia = round(float(horas_declaradas or 0.0) - float(horas_validadas or 0.0), 4)
    except (TypeError, ValueError):
        diferencia = ""

    monto_validado = _monto_validado(resultado)
    monto_observado = round(resultado.monto_total, 2) if resultado.estado in ESTADOS_SIN_SUSTENTO else 0.0

    return {
        "Fila Excel": resultado.fila_excel or "",
        "Empresa": resultado.empresa or "",
        "Empleado Consolidado": resultado.empleado,
        "Empleado Relatorio": resultado.empleado_relatorio or "",
        "DNI": resultado.dni or "",
        "Fecha": _fecha_fmt(resultado.fecha),
        "Turno": resultado.turno or "",
        "Hora Inicio": _hora_fmt(resultado.hora_entrada),
        "Hora Final": _hora_fmt(resultado.hora_salida),
        "Fecha Salida": _fecha_fmt(resultado.fecha_salida),
        "Score Nombre": resultado.score,
        "Confianza": resultado.confianza or "",
        "Metodo Nombre": resultado.metodo_nombre or "",
        "Estado": resultado.estado or "",
        "Observación": resultado.observacion,
        "Horas Declaradas": horas_declaradas if horas_declaradas is not None else "",
        "Horas Validadas": horas_validadas,
        "Horas Trabajadas": resultado.horas_trabajadas or 0.0,
        "Horas Netas (menos comida)": resultado.horas_netas or 0.0,
        "Jornada Nominal": resultado.jornada_nominal
            if resultado.jornada_nominal is not None else "",
        "Horas Extras": _horas_extras_fmt(resultado),
        "Diferencia Horas": diferencia,
        "Monto Declarado": round(resultado.monto_total, 2),
        "Monto Validado": monto_validado,
        "Monto Observado": monto_observado,
        "Ahorro Potencial": monto_observado,
        "Validacion Rainbow": resultado.validacion_rainbow or "",
        "Tipo HHEE": resultado.tipo_hhee or "",
    }


def generar_log(resultados, carpeta_logs, periodo_nombre):
    """Escribe el log de auditoría CSV del período."""
    nombre = "LOG_%s.csv" % periodo_nombre
    ruta = os.path.join(carpeta_logs, nombre)
    with open(ruta, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=COLUMNAS_LOG)
        writer.writeheader()
        for r in resultados:
            writer.writerow(_fila_log(r))
    return ruta


def generar_historico(resultados, productividad, impacto, cfg, periodo_nombre,
                      fecha_ejecucion=None, log_ruta=None, dashboard=None):
    """Contruye el paquete histórico del período (JSON + copia del log)."""
    fecha_ejecucion = fecha_ejecucion or datetime.now()
    anio = fecha_ejecucion.strftime("%Y")
    carpeta = os.path.join(cfg["rutas"]["historicos_dir"], anio, periodo_nombre)
    os.makedirs(carpeta, exist_ok=True)

    resumen = {
        "periodo": periodo_nombre,
        "fecha_ejecucion": fecha_ejecucion.isoformat(),
        "version_sistema": cfg.get("version_sistema"),
        "config": cfg,
        "productividad": productividad,
        "impacto": impacto,
        "registros": [_fila_log(r) for r in resultados],
    }
    if dashboard:
        resumen["dashboard"] = dashboard
    ruta_json = os.path.join(carpeta, "resumen_%s.json" % fecha_ejecucion.strftime("%Y%m%d_%H%M%S"))
    with open(ruta_json, "w", encoding="utf-8") as fh:
        json.dump(resumen, fh, ensure_ascii=False, indent=2, default=str)

    if log_ruta and os.path.exists(log_ruta):
        destino_log = os.path.join(carpeta, os.path.basename(log_ruta))
        with open(log_ruta, encoding="utf-8-sig") as fh_in:
            contenido = fh_in.read()
        with open(destino_log, "w", encoding="utf-8-sig") as fh_out:
            fh_out.write(contenido)
    return carpeta


def _celda_vacia(valor):
    return valor is None or str(valor).strip() == ""


CONVENCION_COLUMNAS = {
    "HORA INICIO": "AZ",
    "HORA FINAL": "BA",
    "Comentarios": "BE",
    "H-H REV": "AC",
    "TIPO": "AG",
    "HORA INICIO - HORA FINAL": "BB",
    "HORAS TRABAJADAS": "BC",
    "HORAS TRABAJADAS HEXAGESIMALES": "BD",
    "25%": "AD",
    "35%": "AE",
    "100%": "AF",
    "Costo al 25%": "AI",
    "Costo al 35%": "AJ",
    "Costo al 100%": "AK",
    "Total HHEE (S/)": "AL",
    "Cantidad Movilidad": "AM",
    "Movilidad": "AN",
    "Total Movil (S/)": "AO",
    "Costo Alimentacion": "AP",
    "Total HHEE + Transp. + Aliment (S/)": "AQ",
}


def _columna_destino(reg, nombre_col):
    """Letra de columna destino: la del archivo si existe, si no la de la
    convención corporativa (AZ/BA/BE), o None si no hay forma de escribirla."""
    return reg["columnas"].get(nombre_col) or CONVENCION_COLUMNAS.get(nombre_col)


def _valor_para_columna(nombre_col, res, cfg):
    """Valor y política de escritura para una columna destino.

    Devuelve (escribir_siempre, valor): si escribir_siempre es True la celda
    se sobrescribe aunque tenga contenido (resultado final del sistema); si es
    False se respeta el llenado manual y solo se escribe sobre celdas vacías.
    valor None indica que esta columna no se escribe para este registro.
    """
    cols = cfg["lectura"]
    if nombre_col == cols["col_hora_inicio"]:
        return (True, _hora_fmt(res.hora_entrada)) if res.hora_entrada else (False, None)
    if nombre_col == cols["col_hora_fin"]:
        return (True, _hora_fmt(res.hora_salida)) if res.hora_salida else (False, None)
    if nombre_col == cols["col_validacion"]:
        vc = cfg.get("salida", {}).get("validacion_rainbow", {})
        return True, (res.validacion_rainbow or vc.get("no", "NO"))
    if nombre_col == cols.get("col_tipo_hhee") and res.tipo_hhee:
        return True, res.tipo_hhee
    if nombre_col == cols["col_hora_inicio_obj"]:
        return (False, _hora_fmt(res.hora_entrada)) if res.hora_entrada else (False, None)
    if nombre_col == cols["col_hora_fin_obj"]:
        return (False, _hora_fmt(res.hora_salida)) if res.hora_salida else (False, None)
    if nombre_col == cols["col_comentarios"]:
        comentario = _escribir_comentario(None, None, res, cfg)
        return False, (comentario or None)
    if nombre_col == cols.get("col_hh_rev"):
        return (True, res.hh_rev) if res.hh_rev is not None else (False, None)
    if nombre_col == cols.get("col_tipo_rev") and res.tipo_rev:
        return True, res.tipo_rev
    if nombre_col == cols.get("col_duracion_txt"):
        return (True, res.duracion_txt) if res.duracion_txt else (False, None)
    if nombre_col == cols.get("col_sin_almuerzo"):
        return (True, res.sin_almuerzo_txt) if res.sin_almuerzo_txt else (False, None)
    if nombre_col == cols.get("col_hexagesimal"):
        return (True, res.hexagesimal) if res.hexagesimal is not None else (False, None)
    # --- Columnas de costos ---
    if nombre_col == "25%":
        return (True, res.horas_25) if res.horas_25 else (False, None)
    if nombre_col == "35%":
        return (True, res.horas_35) if res.horas_35 else (False, None)
    if nombre_col == "100%":
        return (True, res.horas_100) if res.horas_100 else (False, None)
    if nombre_col == "Costo al 25%":
        return (True, res.costo_25) if res.costo_25 else (False, None)
    if nombre_col == "Costo al 35%":
        return (True, res.costo_35) if res.costo_35 else (False, None)
    if nombre_col == "Costo al 100%":
        return (True, res.costo_100) if res.costo_100 else (False, None)
    if nombre_col == "Total HHEE (S/)":
        return (True, res.valor_hhee) if res.valor_hhee else (False, None)
    if nombre_col == "Cantidad Movilidad":
        return (True, res.transporte_cant) if res.transporte_cant else (False, None)
    if nombre_col == "Movilidad":
        return (True, res.transporte_valor) if res.transporte_valor else (False, None)
    if nombre_col == "Total Movil (S/)":
        val = res.transporte_valor if res.transporte_valor else 0.0
        return (True, val) if res.transporte_cant else (False, None)
    if nombre_col == "Costo Alimentacion":
        return (True, res.alimentacion_valor) if res.alimentacion_valor else (False, None)
    if nombre_col == "Total HHEE + Transp. + Aliment (S/)":
        return (True, res.costo_total) if res.costo_total else (False, None)
    return False, None


def _fila_encabezado(registros, defecto=11):
    if not registros:
        return defecto
    return max(1, min(reg["fila_excel"] for reg in registros) - 1)


def _escribir_comentario(hoja, celda, res, cfg):
    """Determina el comentario automático para BE (solo si está vacío).

    Los estados sin evidencia real (ERROR, PENDIENTE, AMBIGUO, REVISIÓN
    MANUAL y OBSERVADO) no reciben un comentario fabricado: se deja la
    celda vacía para la revisión manual.
    """
    if res.estado in ("ERROR", "PENDIENTE", "AMBIGUO", "REVISIÓN MANUAL", "OBSERVADO"):
        return None
    textos = cfg.get("comentarios", {})
    if res.hora_entrada is None and res.hora_salida is None:
        return textos.get("sin_marcacion")
    if res.hora_entrada is None:
        return textos.get("sin_entrada")
    if res.hora_salida is None:
        return textos.get("sin_salida")
    if res.posible_mas12:
        return textos.get("texto_mas12")
    return None


def _col_indice(letra):
    """Convierte letras de columna (AZ, BA...) en índice 1-based."""
    indice = 0
    for c in letra:
        indice = indice * 26 + (ord(c) - ord("A") + 1)
    return indice


def _excel_disponible():
    """True si pywin32 + Microsoft Excel están instalados (escritura fiel)."""
    try:
        import win32com.client  # noqa: F401
        return True
    except Exception:
        return False


def verificar_escritura(ruta, registros, resultados, cfg):
    """Re-abre el consolidado y cuenta celdas objetivo escritas vs esperadas.

    Detecta el fallo silencioso corregido en v0.4.2 (columnas objetivo
    ausentes sin error). Devuelve {"esperadas", "escritas", "faltantes"}
    o {"error": ...} si no se puede re-abrir el archivo.

    Rendimiento v1.1: una sola pasada secuencial (read_only) recolecta las
    celdas objetivo; antes se consultaba celda por celda.
    """
    esperadas = {c: 0 for c in cfg["salida"]["modificar_celdas"]}
    escritas = {c: 0 for c in cfg["salida"]["modificar_celdas"]}
    faltantes = []
    try:
        libro = openpyxl.load_workbook(ruta, data_only=False, read_only=True)
        hoja = libro[cfg["lectura"]["hoja_consolidado"]]
    except Exception:
        return {"error": "no se pudo reabrir el consolidado para verificar (v0.4.4)"}
    try:
        result_por_fila = {r.fila_excel: r for r in resultados}
        cols = cfg["lectura"]
        nombres_rev = {cols.get(k) for k in
                       ("col_hh_rev", "col_tipo_rev", "col_duracion_txt",
                        "col_sin_almuerzo", "col_hexagesimal")}
        nombres_rev.update({"25%", "35%", "100%", "Costo al 25%",
                            "Costo al 35%", "Costo al 100%", "Total HHEE (S/)",
                            "Cantidad Movilidad", "Movilidad", "Total Movil (S/)",
                            "Costo Alimentacion",
                            "Total HHEE + Transp. + Aliment (S/)"})
        # {fila -> {letra -> nombre_col}} objetivo por fila
        objetivos = {}
        for reg in registros:
            res = result_por_fila.get(reg["fila_excel"])
            if res is None:
                continue
            for nombre_col in cfg["salida"]["modificar_celdas"]:
                if nombre_col == cols["col_hora_inicio_obj"]:
                    esperado = _hora_fmt(res.hora_entrada)
                elif nombre_col == cols["col_hora_fin_obj"]:
                    esperado = _hora_fmt(res.hora_salida)
                elif nombre_col in nombres_rev:
                    _siempre, valor = _valor_para_columna(nombre_col, res, cfg)
                    esperado = "" if valor is None else str(valor)
                else:
                    esperado = _escribir_comentario(None, None, res, cfg) or ""
                if not esperado:
                    continue
                letra = _columna_destino(reg, nombre_col)
                if not letra:
                    continue
                esperadas[nombre_col] += 1
                objetivos.setdefault(reg["fila_excel"], {})[letra] = nombre_col

        if objetivos:
            idx_por_letra = {openpyxl.utils.column_index_from_string(l): l
                             for l in {l for m in objetivos.values() for l in m}}
            f_min, f_max = min(objetivos), max(objetivos)
            vistos = {}
            for fila in hoja.iter_rows(min_row=f_min, max_row=f_max):
                n = getattr(fila[0], "row", None)
                obj_fila = objetivos.get(n)
                if not obj_fila:
                    continue
                for celda in fila:
                    letra = idx_por_letra.get(getattr(celda, "column", None))
                    if letra is not None and letra in obj_fila:
                        vistos[(n, letra)] = celda.value
            for fila_n, mapa_fila in objetivos.items():
                for letra, nombre_col in mapa_fila.items():
                    valor = vistos.get((fila_n, letra))
                    if not _celda_vacia(valor):
                        escritas[nombre_col] += 1
                    else:
                        faltantes.append((fila_n, letra, nombre_col))
    finally:
        libro.close()
    return {"esperadas": esperadas, "escritas": escritas,
            "faltantes": faltantes[:20], "total_faltantes": len(faltantes)}


def _com_reintentar(accion, intentos=6, pausa=0.7):
    """Ejecuta una llamada COM a Excel reintentando si está ocupado.

    Cuando Excel tiene un diálogo abierto, arranca o está saturado rechaza las
    llamadas con 'La llamada fue rechazada por el destinatario' (-2147418111)
    o 'El servidor RPC no está disponible' transitorio (-2147417846); en esos
    casos se reintenta con pausa creciente antes de fallar.
    """
    import time
    try:
        import pywintypes
    except Exception:
        return accion()
    for i in range(intentos):
        try:
            return accion()
        except pywintypes.com_error as exc:
            hresult = exc.args[0] if exc.args else None
            if hresult not in (-2147418111, -2147417846) or i == intentos - 1:
                raise
            time.sleep(pausa * (i + 1))


def _escribir_con_excel(ruta_salida, registros, resultados, cfg):
    """Escribe AZ/BA/BE y columnas de plantilla abriendo el archivo CON
    Microsoft Excel (COM).

    A diferencia de openpyxl (que al re-guardar elimina tablas pivote, vínculos
    externos y propiedades del libro, provocando el aviso de "reparar archivo"
    en Excel), esta vía conserva TODO el contenido original: Excel mismo abre
    y guarda el libro.

    Rendimiento v1.1: las decisiones de escritura se calculan en Python y las
    celdas se escriben por TRAMOS contiguos (una llamada COM por tramo) en vez
    de celda por celda; con consolidados grandes reduce el tiempo de escritura
    de ~1 minuto a segundos sin cambiar el resultado celda a celda.
    """
    import shutil
    import win32com.client

    # ---------- plan de escritura calculado en Python (sin llamadas COM)
    result_por_fila = {r.fila_excel: r for r in resultados}
    fila_enc = _fila_encabezado(registros)
    # {letra -> {"col_idx": int, "desde_archivo": bool,
    #            "filas": {fila -> (escribir_siempre, valor)}}}
    planes = {}
    for reg in registros:
        res = result_por_fila.get(reg["fila_excel"])
        if res is None:
            continue
        for nombre_col in cfg["salida"]["modificar_celdas"]:
            letra = _columna_destino(reg, nombre_col)
            if not letra:
                continue
            escribir_siempre, valor = _valor_para_columna(nombre_col, res, cfg)
            if valor is None:
                continue
            plan = planes.setdefault(letra, {
                "col_idx": _col_indice(letra),
                "desde_archivo": bool(reg["columnas"].get(nombre_col)),
                "filas": {}})
            plan["filas"][reg["fila_excel"]] = (escribir_siempre, valor)

    excel = _com_reintentar(lambda: win32com.client.DispatchEx("Excel.Application"))
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    try:
        wb = _com_reintentar(lambda: excel.Workbooks.Open(ruta_salida))
        hoja = _com_reintentar(
            lambda: wb.Worksheets(cfg["lectura"]["hoja_consolidado"]))

        # ---------- escritura por tramos contiguos por columna
        # Rehacemos el recorrido conservando el nombre de columna para los
        # encabezados creados sobre la marcha (convención AZ/BA/BE).
        nombres_por_letra = {}
        for reg in registros:
            res = result_por_fila.get(reg["fila_excel"])
            if res is None:
                continue
            for nombre_col in cfg["salida"]["modificar_celdas"]:
                letra = _columna_destino(reg, nombre_col)
                if letra and not reg["columnas"].get(nombre_col):
                    nombres_por_letra.setdefault(letra, nombre_col)

        # Columna de convención (AZ/BA/BE): encabezado solo si está vacío
        for letra, nombre_col in nombres_por_letra.items():
            col_idx = _col_indice(letra)
            try:
                cab_actual = hoja.Cells(fila_enc, col_idx).Value
            except Exception:
                cab_actual = None
            if _celda_vacia(cab_actual):
                hoja.Cells(fila_enc, col_idx).Value = nombre_col

        for letra, plan in planes.items():
            col_idx = plan["col_idx"]
            filas_plan = plan["filas"]
            r1, r2 = min(filas_plan), max(filas_plan)
            rango = hoja.Range(hoja.Cells(r1, col_idx), hoja.Cells(r2, col_idx))
            vals = _com_reintentar(lambda rg=rango: rg.Value)
            forms = _com_reintentar(lambda rg=rango: rg.Formula)
            if not isinstance(vals, tuple):          # rango de una sola celda
                vals, forms = ((vals,),), ((forms,),)
            vals = [v[0] if isinstance(v, tuple) else v for v in vals]
            forms = [f[0] if isinstance(f, tuple) else f for f in forms]

            inicio_tramo = None                      # fila del primer escrito del tramo
            valores_tramo = []

            def _cerrar_tramo(fin):
                if inicio_tramo is None:
                    return
                tramo = hoja.Range(hoja.Cells(inicio_tramo, col_idx),
                                   hoja.Cells(fin, col_idx))
                if all(isinstance(v, str) for v in valores_tramo):
                    tramo.NumberFormat = "@"
                tramo.Value = tuple((v,) for v in valores_tramo)

            for i, fila in enumerate(range(r1, r2 + 1)):
                decision = filas_plan.get(fila)
                escribe = False
                if decision is not None:
                    escribir_siempre, valor = decision
                    if valor is not None:
                        formula = forms[i]
                        es_formula = isinstance(formula, str) and \
                            formula.strip().startswith("=")
                        if not es_formula and (escribir_siempre or
                                               _celda_vacia(vals[i])):
                            escribe = True
                if escribe:
                    if inicio_tramo is None:
                        inicio_tramo = fila
                    valores_tramo.append(decision[1])
                else:
                    _cerrar_tramo(fila - 1)
                    inicio_tramo = None
                    valores_tramo = []
            _cerrar_tramo(r2)

        try:
            wb.FullCalcOnLoad = True
        except Exception:
            pass
        try:
            _com_reintentar(lambda: setattr(excel, "Calculation", -4105))  # xlCalculationAutomatic
        except Exception:
            pass
        _com_reintentar(lambda: wb.Save())
        _com_reintentar(lambda: wb.Close(False))
        wb = None
    finally:
        if wb is not None:
            try:
                _com_reintentar(lambda: wb.Close(False))
            except Exception:
                pass
        try:
            _com_reintentar(lambda: excel.Quit())
        except Exception:
            # Excel quedó colgado (diálogo modal o instancia huérfana): se
            # termina SOLO el proceso de esta instancia de automatización,
            # nunca una ventana visible del usuario.
            _terminar_excel_colgado(excel)


def _terminar_excel_colgado(excel):
    """Cierra por la fuerza el EXCEL.EXE de esta instancia COM si quedó colgado.

    Solo actúa si la instancia NO tiene ventana visible (título vacío), es
    decir, una instancia de automatización sin trabajo del usuario a la vista.
    """
    try:
        import ctypes
        import win32process
        hwnd = excel.Hwnd
        titulo = ctypes.windll.user32.GetWindowTextW(hwnd) or ""
        if str(titulo).strip():
            return  # hay ventana visible: no se toca
        pid = win32process.GetWindowThreadProcessId(hwnd)[1]
        import subprocess
        subprocess.run(["taskkill", "/PID", str(pid), "/F"], capture_output=True,
                       timeout=15)
    except Exception:
        pass


def escribir_consolidado(consolidado_ruta, registros, resultados, cfg, periodo_nombre):
    """Escribe una copia del consolidado completando solo las celdas objetivo.

    v0.2.0: escribe AZ (HORA INICIO) y BA (HORA FINAL) en texto y BE
    (Comentarios) automáticos, SOLO si las celdas están vacías (nunca
    sobrescribe el llenado manual). BB/BC/BD quedan como fórmulas y el
    libro se guarda con fullCalcOnLoad para que Excel las recalcule.

    v0.5.0: además de AZ/BA/BE, el resultado final se escribe en las columnas
    de la plantilla: Hora de Inicio y Hora Fin (se sobrescriben con el valor
    validado), VALIDACION RAINBOW (SI/NO) y TIPO DE HHEE (clasificación
    automática). AZ/BA siguen como columnas auxiliares y solo se llenan si
    están vacías.

    v0.3.1: si Microsoft Excel está instalado se escribe vía COM, que
    conserva tablas pivote, vínculos externos, imágenes y propiedades del
    libro (el re-guardado con openpyxl provocaba el aviso de "reparar
    archivo" de Excel). openpyxl queda como vía de respaldo sin Excel.
    """
    import shutil

    cols = cfg["lectura"]

    os.makedirs(cfg["rutas"]["output_dir"], exist_ok=True)
    ruta = os.path.join(cfg["rutas"]["output_dir"], cfg["salida"]["nombre_archivo"])
    if os.path.exists(ruta) and not cfg["salida"]["sobrescribir"]:
        ruta = os.path.join(cfg["rutas"]["output_dir"],
                            "CONSOLIDADO_COMPLETADO_%s.xlsx" % periodo_nombre)
    ruta = os.path.abspath(ruta)  # Excel (COM) necesita ruta absoluta

    if _excel_disponible():
        try:
            try:
                shutil.copy2(consolidado_ruta, ruta)
            except PermissionError as exc:
                if os.path.exists(ruta):
                    base, ext = os.path.splitext(ruta)
                    ruta = "%s_%s%s" % (base, datetime.now().strftime("%H%M%S"), ext)
                    try:
                        shutil.copy2(consolidado_ruta, ruta)
                    except PermissionError:
                        raise RuntimeError("No se pudo leer el archivo original. Ciérrelo en Excel e intente nuevamente.")
                else:
                    raise RuntimeError("No se pudo leer el archivo original o la carpeta de destino está protegida.")
            _escribir_con_excel(ruta, registros, resultados, cfg)
            return ruta
        except Exception as exc:
            mensaje = str(exc)
            if "No se pudo leer el archivo original" in mensaje:
                raise RuntimeError(mensaje)
            raise RuntimeError(
                "Excel no pudo guardar el archivo (%s). Verifique que el archivo "
                "no este abierto en otra ventana de Excel y vuelva a intentar. "
                "Archivo solicitado: %s" % (mensaje[:180], ruta))

    libro = openpyxl.load_workbook(consolidado_ruta, data_only=False)
    hoja = libro[cols["hoja_consolidado"]]

    result_por_fila = {r.fila_excel: r for r in resultados}
    fila_enc = _fila_encabezado(registros)
    cabeceras_escritas = set()

    for reg in registros:
        res = result_por_fila.get(reg["fila_excel"])
        if res is None:
            continue
        mapa_col = reg["columnas"]
        for nombre_col in cfg["salida"]["modificar_celdas"]:
            letra = _columna_destino(reg, nombre_col)
            if not letra:
                continue
            if nombre_col not in cabeceras_escritas and not mapa_col.get(nombre_col):
                ref_enc = letra + str(fila_enc)
                if _celda_vacia(hoja[ref_enc].value):
                    hoja[ref_enc] = nombre_col
                cabeceras_escritas.add(nombre_col)
                
            escribir_siempre, valor = _valor_para_columna(nombre_col, res, cfg)
            if valor is None:
                continue
            ref = letra + str(reg["fila_excel"])
            if not escribir_siempre and not _celda_vacia(hoja[ref].value):
                continue
            hoja[ref] = valor

    libro.calculation.fullCalcOnLoad = True
    libro.save(ruta)
    return ruta


def escribir_excel_validacion(ruta_excel, registros, hallazgos_por_fila,
                               resultados, cfg, periodo_nombre="VALIDADO"):
    """Copia el Excel y agrega columnas VALIDACIÓN | OBSERVACIÓN | ESTADO.

    Cada fila recibe el resumen de sus hallazgos de validación.
    Devuelve la ruta del archivo generado.
    """
    import shutil

    cols = cfg["lectura"]
    os.makedirs(cfg["rutas"]["output_dir"], exist_ok=True)
    ruta = os.path.join(cfg["rutas"]["output_dir"],
                        "EXCEL_VALIDADO_%s.xlsx" % periodo_nombre)
    ruta = os.path.abspath(ruta)

    try:
        shutil.copy2(ruta_excel, ruta)
    except PermissionError:
        ruta = os.path.join(cfg["rutas"]["output_dir"],
                            "EXCEL_VALIDADO_%s_%s.xlsx" % (
                                periodo_nombre,
                                datetime.now().strftime("%H%M%S")))
        shutil.copy2(ruta_excel, ruta)

    if _excel_disponible():
        try:
            _escribir_validacion_con_excel(ruta, registros, hallazgos_por_fila,
                                           resultados, cols)
            return ruta
        except Exception:
            pass

    # Fallback openpyxl
    libro = openpyxl.load_workbook(ruta)
    hoja = libro[cols["hoja_consolidado"]]
    fila_enc = _fila_encabezado(registros)

    # Encontrar primera columna vacía después de las existentes
    max_col = 1
    for row in hoja.iter_rows(min_row=fila_enc, max_row=fila_enc):
        for celda in row:
            if celda.column > max_col:
                max_col = celda.column
    # Usar columnas fijas si no se detectan bien
    col_validacion = max_col + 1
    col_observacion = max_col + 2
    col_estado = max_col + 3

    # Escribir encabezados
    hoja.cell(row=fila_enc, column=col_validacion, value="VALIDACIÓN")
    hoja.cell(row=fila_enc, column=col_observacion, value="OBSERVACIÓN")
    hoja.cell(row=fila_enc, column=col_estado, value="ESTADO")

    result_por_fila = {r.fila_excel: r for r in resultados}

    for reg in registros:
        fila = reg["fila_excel"]
        hallazgos = hallazgos_por_fila.get(fila, [])
        if not hallazgos:
            continue

        estado = estado_fila(hallazgos)
        obs = observaciones_fila(hallazgos)

        if estado == "OK":
            texto_validacion = "OK"
            texto_obs = "Información validada correctamente."
        elif estado == "OBSERVADO":
            texto_validacion = "OBSERVADO"
            texto_obs = obs
        else:
            texto_validacion = "ERROR"
            texto_obs = obs

        hoja.cell(row=fila, column=col_validacion, value=texto_validacion)
        hoja.cell(row=fila, column=col_observacion, value=texto_obs)
        hoja.cell(row=fila, column=col_estado, value=estado)

    # Aplicar formato condicional básico (colores)
    try:
        from openpyxl.styles import PatternFill
        verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                            fill_type="solid")
        amarillo = PatternFill(start_color="FFEB9C", end_color="FFEB9C",
                               fill_type="solid")
        rojo = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                           fill_type="solid")
        for reg in registros:
            fila = reg["fila_excel"]
            estado = hoja.cell(row=fila, column=col_estado).value
            if estado == "OK":
                hoja.cell(row=fila, column=col_estado).fill = verde
            elif estado == "OBSERVADO":
                hoja.cell(row=fila, column=col_estado).fill = amarillo
            elif estado == "ERROR":
                hoja.cell(row=fila, column=col_estado).fill = rojo
    except Exception:
        pass

    libro.save(ruta)
    return ruta


def _escribir_validacion_con_excel(ruta, registros, hallazgos_por_fila,
                                    resultados, cols):
    """Escribe validación usando COM de Excel (preserva formato)."""
    import win32com.client
    import pythoncom

    pythoncom.CoInitialize()
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(ruta)
        ws = wb.Sheets(cols["hoja_consolidado"])

        # Encontrar columna máxima
        max_col = ws.UsedRange.Columns.Count
        col_v = max_col + 1
        col_o = max_col + 2
        col_e = max_col + 3

        # Encabezados
        fila_enc = _fila_encabezado(registros)
        ws.Cells(fila_enc, col_v).Value = "VALIDACIÓN"
        ws.Cells(fila_enc, col_o).Value = "OBSERVACIÓN"
        ws.Cells(fila_enc, col_e).Value = "ESTADO"

        result_por_fila = {r.fila_excel: r for r in resultados}

        verde = 0xC6EFCE
        amarillo = 0xFFEB9C
        rojo = 0xFFC7CE

        for reg in registros:
            fila = reg["fila_excel"]
            hallazgos = hallazgos_por_fila.get(fila, [])
            if not hallazgos:
                continue

            estado = estado_fila(hallazgos)
            obs = observaciones_fila(hallazgos)

            if estado == "OK":
                ws.Cells(fila, col_v).Value = "OK"
                ws.Cells(fila, col_o).Value = "Información validada correctamente."
                ws.Cells(fila, col_e).Value = "OK"
                ws.Cells(fila, col_e).Interior.Color = verde
            elif estado == "OBSERVADO":
                ws.Cells(fila, col_v).Value = "OBSERVADO"
                ws.Cells(fila, col_o).Value = obs
                ws.Cells(fila, col_e).Value = "OBSERVADO"
                ws.Cells(fila, col_e).Interior.Color = amarillo
            else:
                ws.Cells(fila, col_v).Value = "ERROR"
                ws.Cells(fila, col_o).Value = obs
                ws.Cells(fila, col_e).Value = "ERROR"
                ws.Cells(fila, col_e).Interior.Color = rojo

        wb.Save()
        wb.Close()
    finally:
        excel.Quit()
        pythoncom.CoUninitialize()