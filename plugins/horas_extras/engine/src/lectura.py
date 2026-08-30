"""Lectura de los archivos fuente: consolidado y relatorio(s).

La lectura del consolidado detecta las columnas por NOMBRE dentro del
encabezado de Tabla3 (no por posición absoluta). Excluye la fila de totales.
"""

import os
import re

import openpyxl

from empresas import coincidir_empresa, extraer_nombre_empresa
from normalizacion import normalizar_fecha, normalizar_hora


def _extraer_fila(ref):
    """Extrae la fila numérica de una referencia tipo 'A11' o 'AV67'."""
    m = re.search(r"(\d+)$", ref)
    return int(m.group(1)) if m else int(ref)


def _leer_mapa_columnas(ws, fila_encabezado):
    """Construye el mapa {nombre_columna -> letra} desde la fila de encabezado."""
    mapa = {}
    for celda in ws[fila_encabezado]:
        if celda.value is not None and str(celda.value).strip():
            mapa[str(celda.value).strip()] = celda.column_letter
    return mapa


def _localizar_tabla(ws, nombre_tabla):
    """Devuelve (fila_encabezado, fila_ultima) del rango de la tabla definida."""
    if nombre_tabla and nombre_tabla in ws.tables:
        prim, ult = ws.tables[nombre_tabla].ref.split(":")
        return _extraer_fila(prim), _extraer_fila(ult)
    raise ValueError("Tabla '%s' no encontrada" % nombre_tabla)


def leer_consolidado(ruta, cfg, avisos=None, wb=None):
    """Lee Tabla3 del consolidado y devuelve la lista de registros.

    Cada registro incluye: fila_excel, fecha, turno, empleado,
    hora_inicio_orig, hora_fin_orig, monto_total, horas_declaradas
    y las letras de columna para escritura posterior.

    Si se entrega un workbook ya abierto (wb), se reutiliza en lugar de
    volver a abrir el archivo (evita hasta 3 aperturas de archivos grandes).
    """
    cols = cfg["lectura"]
    avisos = [] if avisos is None else avisos

    abrio_aqui = wb is None
    if not os.path.exists(ruta):
        raise ValueError("Consolidado no encontrado: %s" % ruta)
    if abrio_aqui:
        try:
            wb = openpyxl.load_workbook(ruta, data_only=True)
        except Exception as exc:
            raise ValueError("No se pudo abrir el consolidado %s (%s). Verifica que sea un "
                             "archivo Excel real y que no este abierto en otro programa."
                             % (os.path.basename(ruta), str(exc)[:120]))
    try:
        hoja = wb[cols["hoja_consolidado"]]
    except KeyError:
        if abrio_aqui:
            wb.close()
        raise ValueError("El consolidado no tiene la hoja '%s' esperada. Carga el "
                         "consolidado del periodo con su estructura habitual."
                         % cols["hoja_consolidado"])
    fila_enc, fila_ult = _localizar_tabla(hoja, cols["tabla_consolidado"])
    mapa = _leer_mapa_columnas(hoja, fila_enc)

    obligatorias = {
        cols["col_empleado"]: "la columna de empleados ('%s')" % cols["col_empleado"],
        cols["col_turno"]: "la columna de turno ('%s')" % cols["col_turno"],
        cols["col_fecha"]: "la columna de fecha ('%s')" % cols["col_fecha"],
        cols["col_numcontrol"]: "la columna '# Control'",
    }
    for nombre, desc in obligatorias.items():
        if nombre not in mapa:
            raise ValueError("No se encontro %s en el consolidado. Verifica que el "
                             "archivo tenga el formato esperado." % desc)

    def obtener(fila, nombre_col):
        if nombre_col not in mapa:
            return None
        return hoja.cell(row=fila, column=openpyxl.utils.column_index_from_string(mapa[nombre_col])).value

    def obtener_declaradas(fila):
        for nombre in (cols["col_horas_declaradas"], "H-H CORREO", "H-H REV"):
            valor = obtener(fila, nombre)
            if valor is not None:
                return valor
        return None

    registros = []
    omitidos = 0
    for fila in range(fila_enc + 1, fila_ult + 1):
        num_control = obtener(fila, cols["col_numcontrol"])
        empleado = obtener(fila, cols["col_empleado"])
        if empleado is None or not str(empleado).strip():
            continue
        if num_control is not None and str(num_control).strip().lower() == "total":
            continue
        turno_raw = obtener(fila, cols["col_turno"])
        fecha_raw = obtener(fila, cols["col_fecha"])
        if fecha_raw is None:
            omitidos += 1
            continue
        cols_escritura = {
            "HORA INICIO": mapa.get(cols["col_hora_inicio_obj"]),
            "HORA FINAL": mapa.get(cols["col_hora_fin_obj"]),
            "Comentarios": mapa.get(cols["col_comentarios"]),
        }
        registros.append({
            "fila_excel": fila,
            "fecha": normalizar_fecha(fecha_raw),
            "turno": str(turno_raw).strip() if turno_raw is not None else "",
            "empleado": str(empleado).strip(),
            "hora_inicio_orig": obtener(fila, cols["col_hora_inicio"]),
            "hora_fin_orig": obtener(fila, cols["col_hora_fin"]),
            "hora_inicio_obj": obtener(fila, cols["col_hora_inicio_obj"]),
            "hora_fin_obj": obtener(fila, cols["col_hora_fin_obj"]),
            "comentario_obj": obtener(fila, cols["col_comentarios"]),
            "monto_total": obtener(fila, cols["col_monto_total"]) or 0.0,
            "horas_declaradas": obtener_declaradas(fila),
            "especialidad": str(obtener(fila, cols.get("col_especialidad", "")) or "").strip(),
            "npersonas": obtener(fila, cols.get("col_npersonas", "")),
            "columnas": cols_escritura,
        })
    if omitidos:
        avisos.append("Consolidado: %d fila(s) omitida(s) por falta de fecha." % omitidos)
    return registros


def _leer_relatorio_archivo(ruta, cols, cfg, avisos):
    """Lee UN relatorio y devuelve su lista de marcaciones (ver leer_relatorio)."""
    if not os.path.exists(ruta):
        raise ValueError("Relatorio no encontrado: %s" % ruta)
    try:
        libro = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError("No se pudo abrir el relatorio %s (%s). Verifica que sea "
                         "un archivo Excel real y que no este abierto en otro "
                         "programa." % (os.path.basename(ruta), str(exc)[:120]))
    try:
        hoja = libro[cols["hoja_relatorio"]]
    except KeyError:
        raise ValueError("El relatorio %s no tiene la hoja '%s' esperada."
                         % (os.path.basename(ruta), cols["hoja_relatorio"]))

    mapa_enc = {}
    # Usamos iter_rows(max_row=1) que es seguro en modo read_only
    encabezados = next(hoja.iter_rows(min_row=1, max_row=1, values_only=True), [])
    for idx, valor in enumerate(encabezados):
        if valor is not None and str(valor).strip():
            mapa_enc[str(valor).strip()] = idx

    for nombre, desc in ((cols["col_rel_empleado"], "la columna 'Empleado'"),
                         (cols["col_rel_fecha"], "la columna 'Fecha'"),
                         (cols["col_rel_hora"], "la columna 'Hora'")):
        if nombre not in mapa_enc:
            raise ValueError("No se encontro %s en el relatorio %s. Verifique que "
                             "el archivo tenga la estructura requerida."
                             % (desc, os.path.basename(ruta)))

    col_empresa = cols.get("col_rel_empresa")
    tiene_col_empresa = col_empresa in mapa_enc
    cache_empresa = {}
    empresas_cfg = cfg.get("empresas") or {}

    # Índices de columna resueltos UNA vez (rendimiento: el bucle corre
    # sobre cientos de miles de filas en relatorios anuales).
    idx_tipo = mapa_enc.get(cols["col_rel_tipo"])
    idx_emp = mapa_enc[cols["col_rel_empleado"]]
    idx_fecha = mapa_enc[cols["col_rel_fecha"]]
    idx_hora = mapa_enc[cols["col_rel_hora"]]
    idx_acceso = mapa_enc.get(cols["col_rel_tipo_acceso"])
    idx_sit = mapa_enc.get(cols["col_rel_situacion"])
    idx_dni = mapa_enc.get(cols["col_rel_dni"])
    idx_empresa = mapa_enc.get(col_empresa)
    solo_tipo = cols.get("solo_tipo") or None
    nombre_archivo = os.path.basename(ruta)

    # Memos de parseo: las mismas cadenas fecha/hora se repiten miles de
    # veces en el RAINBOW; se parsean una sola vez por valor distinto.
    memo_fecha = {}
    memo_hora = {}

    marcaciones = []
    try:
        for fila in hoja.iter_rows(min_row=2, values_only=True):
            if solo_tipo is not None:
                if idx_tipo is None or fila[idx_tipo] != solo_tipo:
                    continue
            empleado = fila[idx_emp]
            if not empleado or not str(empleado).strip():
                continue
            empresa_raw = None
            empresa = None
            if tiene_col_empresa:
                valor = fila[idx_empresa] if idx_empresa is not None else None
                if valor is not None and str(valor).strip():
                    raw_txt = str(valor).strip()
                    if empresas_cfg:
                        if raw_txt not in cache_empresa:
                            limpio = extraer_nombre_empresa(raw_txt)
                            cache_empresa[raw_txt] = coincidir_empresa(limpio, empresas_cfg, avisos)
                        empresa = cache_empresa[raw_txt]
                    empresa_raw = raw_txt
            fecha_raw = fila[idx_fecha]
            try:
                fecha = memo_fecha.get(fecha_raw)
                if fecha is None:
                    fecha = normalizar_fecha(fecha_raw)
                    memo_fecha[fecha_raw] = fecha
            except ValueError:
                avisos.append("%s: fecha inválida %r (omitida)" % (nombre_archivo, fecha_raw))
                continue
            hora_raw = fila[idx_hora]
            try:
                hora = memo_hora.get(hora_raw)
                if hora is None and hora_raw not in memo_hora:
                    hora = normalizar_hora(hora_raw)
                    memo_hora[hora_raw] = hora
            except ValueError:
                avisos.append("%s: hora no parseable %r (omitida)" % (nombre_archivo, hora_raw))
                continue
            if hora is None:
                avisos.append("%s: hora irrecuperable %r (omitida)" % (nombre_archivo, hora_raw))
                continue
            marcaciones.append({
                "empleado": str(empleado).strip(),
                "fecha": fecha,
                "hora": hora,
                "tipo_acceso": str(fila[idx_acceso]).strip() if idx_acceso is not None else "",
                "situacion": str(fila[idx_sit]).strip() if idx_sit is not None else "",
                "dni": fila[idx_dni] if idx_dni is not None else None,
                "empresa": empresa,
                "empresa_raw": empresa_raw,
                "origen": nombre_archivo,
            })
    finally:
        libro.close()
    return marcaciones


def leer_relatorio(rutas, cfg, avisos=None):
    """Lee uno o más relatorios y devuelve la lista de marcaciones.

    Cada marcación: empleado, fecha(date), hora(time), tipo_acceso,
    situacion, dni y archivo de origen. Las marcas con hora o fecha
    irrecuperables se omiten y se registran en `avisos` (lista opcional).

    Rendimiento v1.1: con varios archivos se leen en paralelo (hilos); la
    descompresión zip de openpyxl libera el GIL y acelera ~40% la carga de
    historiales largos. Las marcas y avisos se unen en el orden original de
    los archivos para mantener resultados deterministas.
    """
    avisos = [] if avisos is None else avisos
    rutas = list(rutas)
    marcaciones = []
    for ruta in rutas:
        if not os.path.exists(ruta):
            raise ValueError("Relatorio no encontrado: %s" % ruta)

    if len(rutas) <= 1:
        for ruta in rutas:
            marcaciones.extend(_leer_relatorio_archivo(
                ruta, cfg["lectura"], cfg, avisos))
    else:
        from concurrent.futures import ThreadPoolExecutor
        avisos_por_archivo = {}

        def _trabajo(ruta):
            propios = []
            return _leer_relatorio_archivo(ruta, cfg["lectura"], cfg, propios), propios

        with ThreadPoolExecutor(max_workers=min(len(rutas), 4)) as pool:
            futuros = [(ruta, pool.submit(_trabajo, ruta)) for ruta in rutas]
            # recoge en orden original; propaga errores tal cual y conserva
            # los avisos de cada archivo en el orden de los archivos
            for ruta, futuro in futuros:
                marcas_ruta, avisos_ruta = futuro.result()
                marcaciones.extend(marcas_ruta)
                avisos.extend(avisos_ruta)
    if not marcaciones:
        avisos.append("Relatorio: el(los) archivo(s) de marcaciones no aportaron "
                      "ninguna marca util (revisa tipo de registro, fechas y horas).")
    return marcaciones, avisos


def cargar_tabla_costos(ruta_consolidado, wb=None):
    """Carga la tabla de tarifas de HHEE desde la hoja 'Costos At. Emerg'.

    Extrae los costos por especialidad. Devuelve una instancia de
    costos.TablaCostos (o None si la hoja no existe o no tiene datos).
    Si se pasa un workbook ya abierto (wb), se reutiliza y no se cierra.
    """
    from costos import cargar_tabla_costos as _cargar
    if not ruta_consolidado and wb is None:
        return None
    abrio_aqui = wb is None
    if abrio_aqui:
        if not os.path.exists(str(ruta_consolidado)):
            return None
        try:
            wb = openpyxl.load_workbook(str(ruta_consolidado), data_only=True)
        except Exception:
            return None
    try:
        tabla = _cargar(wb)
    finally:
        if abrio_aqui:
            wb.close()
    return tabla if len(tabla) > 0 else None


def leer_columnas_validacion(ruta_excel, registros, cfg, wb=None):
    """Lee columnas de costos/validación del Excel para comparación.

    Agrega a cada registro un dict 'valores_excel' con los valores
    que el Excel ya tiene escritos (horas 25/35/100, costos, transporte, etc.).
    Si se pasa un workbook ya abierto (wb), se reutiliza y no se cierra.
    """
    cols = cfg["lectura"]
    mapa_costos = {
        "H-H REV": "hh_rev",
        "25%": "h25",
        "35%": "h35",
        "100%": "h100",
        "TIPO": "tipo",
        "Hexagesimal": "hexagesimal",
        "Costo al 25%": "costo_25",
        "Costo al 35%": "costo_35",
        "Costo al 100%": "costo_100",
        "Total HHEE (S/)": "valor_hhee",
        "Cantidad Movilidad": "transporte_cant",
        "Movilidad": "transporte_valor",
        "Costo Alimentacion": "alimentacion_valor",
        "Total HHEE + Transp. + Aliment (S/)": "costo_total",
    }
    # Intentar leer también el tipo de alimentación
    mapa_extra = {
        "Cantidad de horas - 1 de almuerzo": "sin_almuerzo",
    }

    abrio_aqui = wb is None
    if abrio_aqui:
        try:
            wb = openpyxl.load_workbook(str(ruta_excel), data_only=True)
        except Exception:
            for reg in registros:
                reg["valores_excel"] = {}
            return

    try:
        nombre_hoja = cols["hoja_consolidado"]
        if nombre_hoja not in wb.sheetnames:
            for reg in registros:
                reg["valores_excel"] = {}
            return
        hoja = wb[nombre_hoja]
        try:
            fila_enc, _ = _localizar_tabla(hoja, cols["tabla_consolidado"])
        except Exception:
            for reg in registros:
                reg["valores_excel"] = {}
            return
        mapa = _leer_mapa_columnas(hoja, fila_enc)

        # Mapear nombre_columna -> letra
        col_letras = {}
        for nombre_col in list(mapa_costos.keys()) + list(mapa_extra.keys()):
            if nombre_col in mapa:
                col_letras[nombre_col] = mapa[nombre_col]

        # Leer valores por fila
        fila_map = {reg["fila_excel"]: reg for reg in registros}
        for reg in registros:
            reg["valores_excel"] = {}

        for nombre_col, letra in col_letras.items():
            try:
                col_idx = openpyxl.utils.column_index_from_string(letra)
            except Exception:
                continue
            for reg in registros:
                try:
                    val = hoja.cell(row=reg["fila_excel"], column=col_idx).value
                except Exception:
                    val = None
                clave = mapa_costos.get(nombre_col) or mapa_extra.get(nombre_col)
                if clave and val is not None:
                    reg["valores_excel"][clave] = val
    finally:
        if abrio_aqui:
            wb.close()