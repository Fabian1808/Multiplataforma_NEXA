"""Capa de EXPORTACIÓN — Excel multi-hoja.

Genera un workbook con las 6 hojas definidas en config.exportacion.hojas:
  RESUMEN, DETALLE, AUDITORIA, ERRORES, TARIFAS, MARCACIONES.

Con encabezado coloreado (paleta NEXA), filtros y anchos automáticos.
Los montos se escriben como Decimal exacto (2 decimales).
"""

from __future__ import annotations

import os
from datetime import date, datetime
from decimal import Decimal

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import ingesta as ing
import utiles as u


ACCENT = "FF5503"
HEADER_FILL = "FF5503"
HEADER_FONT = "FFFFFF"
ALT_FILL = "FFF3EC"
THIN = Side(style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
_ALT_FILL_OBJ = PatternFill("solid", fgColor=ALT_FILL)  # reutilizable (evita recrear por fila)


def _estilo_encabezado(ws, ncols: int, fila: int = 1) -> None:
    fill = PatternFill("solid", fgColor=HEADER_FILL)
    for c in range(1, ncols + 1):
        cel = ws.cell(row=fila, column=c)
        cel.fill = fill
        cel.font = Font(bold=True, color=HEADER_FONT)
        cel.alignment = Alignment(horizontal="center", vertical="center")
        cel.border = BORDER


def _autoancho(ws, ncols: int, extra: int = 2) -> None:
    """Calcula el ancho de columnas en UNA pasada (no re-itera celdas)."""
    maxlen = [0] * (ncols + 1)
    for fila in ws.iter_rows(min_row=1, max_col=ncols):
        for cel in fila:
            v = cel.value
            if v is not None:
                ln = len(str(v))
                if ln > maxlen[cel.column]:
                    maxlen[cel.column] = ln
    for c in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = min(maxlen[c] + extra, 60)


def _escribir_filas(ws, headers: list, filas: list, inicio: int = 2) -> None:
    """Escribe filas con borde y relleno alternado, reutilizando el fill.

    Si hay muchas filas (reporte grande), se omite el estilo por celda para
    evitar lentitud/memoria excesiva al escribir el Excel.
    """
    estilizar = len(filas) <= 20000
    for i, fila in enumerate(filas, start=inicio):
        alternar = estilizar and i % 2 == 0
        for j, valor in enumerate(fila, start=1):
            cel = ws.cell(row=i, column=j, value=valor)
            if estilizar:
                cel.border = BORDER
                if alternar:
                    cel.fill = _ALT_FILL_OBJ


def _mon(dec, cfg):
    return u.monto_decimal(dec) if dec is not None else Decimal("0.00")


def exportar(resultado, ruta_destino: str | os.PathLike, cfg) -> str:
    wb = openpyxl.Workbook()
    cfg_exp = cfg.get("exportacion") or {}
    hojas = cfg_exp.get("hojas", ["RESUMEN", "DETALLE", "AUDITORIA", "ERRORES", "TARIFAS", "MARCACIONES"])

    # ------------------------------------------------------------------ RESUMEN
    ws = wb.active
    ws.title = hojas[0]
    flias = [
        ["Concepto", "Valor"],
        ["Período (RAINBOW)", "%s — %s" % (_min_fecha(resultado), _max_fecha(resultado))],
        ["Marcaciones RAINBOW", resultado.totales["marcaciones"]],
        ["Personal RELATORIO (maestro)", resultado.totales["empleados"]],
        ["Tarifas (TARIFAS)", resultado.totales["tarifas"]],
        ["Marcaciones conciliadas", resultado.totales["conciliados"]],
        ["Sin conciliar", resultado.totales["sin_conciliar"]],
        ["Jornadas procesadas", resultado.totales["jornadas"]],
        ["Registros de detalle", resultado.totales["registros_detalle"]],
        ["Horas extras totales (h)", resultado.horas_extra_total],
        ["Monto total (HH.EE.)", _mon(resultado.monto_total, cfg)],
    ]
    for estado in resultado.estados:
        flias.append(["Estado %s" % estado, resultado.estados[estado]])
    for i, fila in enumerate(flias, start=1):
        for j, v in enumerate(fila, start=1):
            ws.cell(row=i, column=j, value=v)
    _estilo_encabezado(ws, 2)
    _autoancho(ws, 2)

    # ------------------------------------------------------------------ DETALLE
    ws = wb.create_sheet(hojas[1])
    headers_det = [
        "Fecha", "Empleado", "DNI", "Fotocheck", "Empresa", "RUC", "Cargo",
        "Contrato", "Turno", "Entrada", "Salida", "Horas Trabajadas",
        "Jornada", "Horas Extras", "Tipo Hora", "Horas Tipo", "Tarifa",
        "Monto (S/)", "Nivel Tarifa", "Estado",
    ]
    ws.append(headers_det)
    filas_det = []
    for f in resultado.filas:
        filas_det.append([
            _fmt_fecha(f["fecha"]),
            _limpiar_nombre(f["empleado"]),
            f["dni"], f["fotocheck"],
            f["empresa"], f["ruc"], f["cargo"], f["contrato"],
            f["turno"], f["inicio"], f["fin"],
            float(f["horas_trabajadas"]),
            float(f["jornada"]),
            float(f["horas_extras"]),
            f["tipo_hora"], float(f.get("horas_tipo", 0) or 0),
            float(_mon(f.get("tarifa"), cfg)),
            float(_mon(f.get("monto"), cfg)),
            f["nivel_tarifa"], f["estado"],
        ])
    _escribir_filas(ws, headers_det, filas_det)
    _estilo_encabezado(ws, len(headers_det))
    _autoancho(ws, len(headers_det))
    ws.auto_filter.ref = ws.dimensions

    # ------------------------------------------------------------------ AUDITORIA
    ws = wb.create_sheet(hojas[2])
    headers_aud = ["Fecha", "Empleado", "DNI", "Empresa", "RUC", "Cargo", "Turno",
                   "Metodo conciliación", "Confianza conciliación", "Nivel tarifa",
                   "Confianza tarifa", "Especificación HE", "Monto (S/)"]
    ws.append(headers_aud)
    por_emp = {}
    for rc in resultado.conciliados:
        if not rc.conciliado:
            continue
        emp = rc.empleado
        por_emp.setdefault(emp.empleado, rc)
    filas_aud = []
    vistos = set()
    for f in resultado.filas:
        clave = f["empleado"]
        rc = por_emp.get(clave)
        if clave in vistos:
            continue
        vistos.add(clave)
        filas_aud.append([
            _fmt_fecha(f["fecha"]), _limpiar_nombre(f["empleado"]), f["dni"],
            f["empresa"], f["ruc"], f["cargo"], f["turno"],
            (rc.metodo if rc else "—"),
            ("%.0f" % rc.confianza if rc else "—"),
            f["nivel_tarifa"], ("%.0f" % f["confianza_tarifa"]),
            f.get("especifica", ""), float(_mon(f.get("monto"), cfg)),
        ])
    _escribir_filas(ws, headers_aud, filas_aud)
    _estilo_encabezado(ws, len(headers_aud))
    _autoancho(ws, len(headers_aud))

    # ------------------------------------------------------------------ ERRORES
    ws = wb.create_sheet(hojas[3])
    headers_err = ["Fecha", "Empleado", "DNI", "Cargo", "Tipo", "Detalle"]
    ws.append(headers_err)
    filas_err = []
    for f in resultado.filas:
        if f["estado"] not in ("ERROR", "REVISAR", "ADVERTENCIA"):
            continue
        filas_err.append([
            _fmt_fecha(f["fecha"]), _limpiar_nombre(f["empleado"]), f["dni"],
            f["cargo"], f["estado"], _detalle_error(f),
        ])
    _escribir_filas(ws, headers_err, filas_err)
    _estilo_encabezado(ws, len(headers_err))
    _autoancho(ws, len(headers_err))

    # ------------------------------------------------------------------ TARIFAS
    ws = wb.create_sheet(hojas[4])
    headers_tar = ["Cargo", "Empresa", "RUC", "Objeto del contrato", "25%", "35%", "100%"]
    ws.append(headers_tar)
    filas_tar = []
    for t in resultado.tarifas:
        filas_tar.append([
            t.cargo, t.empresa, t.ruc, t.objeto,
            float(t.c25), float(t.c35), float(t.c100),
        ])
    _escribir_filas(ws, headers_tar, filas_tar)
    _estilo_encabezado(ws, len(headers_tar))
    _autoancho(ws, len(headers_tar))

    # ------------------------------------------------------------------ MARCACIONES
    ws = wb.create_sheet(hojas[5])
    headers_mar = ["Fecha", "Hora", "Tipo Acceso", "Empleado", "DNI", "Fotocheck",
                   "Empresa", "RUC", "Situación", "Método conciliación"]
    ws.append(headers_mar)
    filas_mar = []
    met_por_marc = {}
    for rc in resultado.conciliados:
        met_por_marc[id(rc.marcacion)] = rc.metodo
    for mar in resultado.marcaciones:
        filas_mar.append([
            _fmt_fecha(mar.fecha), mar.hora.strftime("%H:%M:%S") if mar.hora else "",
            mar.tipo_acceso, _limpiar_nombre(mar.empleado), mar.dni, mar.fotocheck,
            mar.empresa, mar.ruc, mar.situacion, met_por_marc.get(id(mar), ""),
        ])
    _escribir_filas(ws, headers_mar, filas_mar)
    _estilo_encabezado(ws, len(headers_mar))
    _autoancho(ws, len(headers_mar))
    ws.auto_filter.ref = ws.dimensions

    os.makedirs(os.path.dirname(os.path.abspath(ruta_destino)) or ".", exist_ok=True)
    wb.save(ruta_destino)
    return str(ruta_destino)


# ---------------------------------------------------------------------------
def _fmt_fecha(f):
    if isinstance(f, (date, datetime)):
        return f.strftime("%d/%m/%Y")
    return f


def _limpiar_nombre(nombre):
    return u.limpiar_codigo(str(nombre)).strip() if nombre else ""


def _min_fecha(res):
    fechas = [m.fecha for m in res.marcaciones if m.fecha]
    return _fmt_fecha(min(fechas)) if fechas else "—"


def _max_fecha(res):
    fechas = [m.fecha for m in res.marcaciones if m.fecha]
    return _fmt_fecha(max(fechas)) if fechas else "—"


def _detalle_error(f):
    if f["estado"] == "ERROR":
        return "Horas extras sin tarifa aplicable (sin match en TARIFAS)."
    if f["estado"] == "REVISAR":
        if f.get("nivel_tarifa") in ("BAJA", "MEDIA"):
            return "Tarifa %s/ambigua — REVISAR manualmente." % f.get("nivel_tarifa")
        return "Revisión requerida."
    return "Horas extras anómalas." if f["estado"] == "ADVERTENCIA" else ""
