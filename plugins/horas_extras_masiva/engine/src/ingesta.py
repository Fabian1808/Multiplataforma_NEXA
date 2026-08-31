"""Capa 1 — INGESTA: lectura de archivos fuente.

Lee RAINBOW (marcaciones), RELATORIO (maestro de personal), TARIFAS,
ÁREAS y GERENCIA detectando columnas por NOMBRE (configuración), no por
posición absoluta. Nunca inventa columnas: si falta una columna requerida
se registra un aviso/error.
"""

from __future__ import annotations

import os
from dataclasses import dataclass, field
from datetime import date, datetime, time
from decimal import Decimal
from typing import Any

import openpyxl

import datos_embebidos as _embed

from utiles import (
    normalizar_fecha, normalizar_hora, normalizar_texto, normalizar_dni,
    limpiar_codigo, limpiar_ruc,
)


class IngestaError(Exception):
    """Error de lectura de un archivo fuente."""


# ---------------------------------------------------------------------------
# Helpers de hoja/encabezados
# ---------------------------------------------------------------------------
def _elegir_hoja(wb, nombre: str | None):
    if nombre:
        if nombre not in wb.sheetnames:
            raise IngestaError("El Excel no tiene la hoja '%s'. Hojas: %s"
                               % (nombre, ", ".join(wb.sheetnames)))
        return wb[nombre]
    # primera hoja no vacía
    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.max_row and ws.max_column:
            return ws
    raise IngestaError("El Excel no tiene hojas con datos.")


def _mapa_encabezados(ws, fila_encabezado: int) -> dict[str, int]:
    mapa = {}
    fila = ws[fila_encabezado]
    for idx, celda in enumerate(fila):
        v = celda.value
        if v is not None and str(v).strip():
            clave = normalizar_texto(v)
            mapa.setdefault(clave, idx)
    return mapa


def _obtener(fila, col_idx):
    if col_idx is None:
        return None
    try:
        return fila[col_idx]
    except IndexError:
        return None


def _col(mapa, cfg, *nombres):
    """Devuelve índice de columna por nombre (primer nombre que aparezca)."""
    for n in nombres:
        idx = mapa.get(normalizar_texto(n))
        if idx is not None:
            return idx
    return None


# ---------------------------------------------------------------------------
# RAINBOW
# ---------------------------------------------------------------------------
@dataclass
class Marcacion:
    tipo: str = ""
    empresa: str = ""
    ruc: str = ""
    num_personal: str = ""
    fotocheck: str = ""
    empleado: str = ""
    dni: str = ""
    fecha: date | None = None
    hora: time | None = None
    tipo_acceso: str = ""
    situacion: str = ""
    origen: str = ""
    fecha_hora: datetime | None = None

    @property
    def es_entrada(self) -> bool:
        return "entrada" in self.tipo_acceso.lower() or "ingreso" in self.tipo_acceso.lower()

    @property
    def es_salida(self) -> bool:
        return "salida" in self.tipo_acceso.lower()

    def clave_dt(self) -> datetime | None:
        if self.fecha and self.hora:
            return datetime.combine(self.fecha, self.hora)
        return self.fecha_hora


def leer_rainbow(ruta, cfg, avisos=None) -> list[Marcacion]:
    """Lee marcaciones RAINBOW desde una ruta o una lista de rutas (varios excels).

    Si `ruta` es una lista/tupla, se procesan todos los archivos y se devuelve
    la concatenación de sus marcaciones.
    """
    avisos = [] if avisos is None else avisos
    rutas = ruta if isinstance(ruta, (list, tuple)) else [ruta]
    if not rutas:
        raise IngestaError("No se indicó ningún archivo RAINBOW.")
    out: list[Marcacion] = []
    for r in rutas:
        out.extend(_leer_rainbow_unico(r, cfg, avisos))
    return out


def _leer_rainbow_unico(ruta, cfg, avisos=None) -> list[Marcacion]:
    avisos = [] if avisos is None else avisos
    if not os.path.exists(ruta):
        raise IngestaError("Rainbow no encontrado: %s" % ruta)
    rc = cfg.get("rainbow") or {}
    try:
        wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    except Exception as exc:
        raise IngestaError("No se pudo abrir Rainbow %s (%s). Verifica que sea un Excel real y que no esté abierto en otro programa."
                           % (os.path.basename(ruta), str(exc)[:120]))
    try:
        ws = _elegir_hoja(wb, rc.get("hoja"))
        fila_enc = int(rc.get("fila_encabezado", 1))
        mapa = _mapa_encabezados(ws, fila_enc)
        idx = {
            "tipo": _col(mapa, rc, rc.get("col_tipo")),
            "empresa": _col(mapa, rc, rc.get("col_empresa")),
            "ruc": _col(mapa, rc, rc.get("col_ruc")),
            "num_personal": _col(mapa, rc, rc.get("col_num_personal")),
            "fotocheck": _col(mapa, rc, rc.get("col_fotocheck")),
            "empleado": _col(mapa, rc, rc.get("col_empleado")),
            "dni": _col(mapa, rc, rc.get("col_dni")),
            "fecha": _col(mapa, rc, rc.get("col_fecha")),
            "hora": _col(mapa, rc, rc.get("col_hora")),
            "tipo_acceso": _col(mapa, rc, rc.get("col_tipo_acceso")),
            "situacion": _col(mapa, rc, rc.get("col_situacion")),
        }
        if idx["empleado"] is None:
            raise IngestaError("Rainbow %s no tiene la columna '%s'. Verifica el formato."
                               % (os.path.basename(ruta), rc.get("col_empleado")))
        permitido = rc.get("permitido_si")
        incluir_denegados = cfg.get("marcaciones", {}).get("incluir_denegados", True)
        out: list[Marcacion] = []
        for fila in ws.iter_rows(min_row=fila_enc + 1, values_only=True):
            empleado = _obtener(fila, idx["empleado"])
            if not empleado or not str(empleado).strip():
                continue
            sit = str(_obtener(fila, idx["situacion"]) or "").strip()
            if permitido and sit and sit != permitido and not incluir_denegados:
                continue
            fecha = _obtener(fila, idx["fecha"])
            hora = _obtener(fila, idx["hora"])
            fechadt = normalizar_fecha(fecha)
            horat = normalizar_hora(hora)
            if fechadt is None or horat is None:
                continue
            mar = Marcacion(
                tipo=str(_obtener(fila, idx["tipo"]) or "").strip(),
                empresa=str(_obtener(fila, idx["empresa"]) or "").strip(),
                ruc=limpiar_ruc(_obtener(fila, idx["ruc"])),
                num_personal=limpiar_codigo(_obtener(fila, idx["num_personal"])),
                fotocheck=limpiar_codigo(_obtener(fila, idx["fotocheck"])),
                empleado=str(empleado).strip(),
                dni=normalizar_dni(_obtener(fila, idx["dni"])),
                fecha=fechadt,
                hora=horat,
                tipo_acceso=str(_obtener(fila, idx["tipo_acceso"]) or "").strip(),
                situacion=sit,
                origen=os.path.basename(ruta),
                fecha_hora=datetime.combine(fechadt, horat),
            )
            out.append(mar)
    finally:
        wb.close()
    return out


# ---------------------------------------------------------------------------
# RELATORIO (maestro de personal)
# ---------------------------------------------------------------------------
@dataclass
class Empleado:
    empresa: str = ""
    unidad: str = ""
    grupo_empresa: str = ""
    codigo_empresa: str = ""
    empresa_terceros: str = ""
    ruc: str = ""
    nombre_comercial: str = ""
    grupo_terceros: str = ""
    empleado: str = ""
    fotocheck: str = ""
    fecha_admision: date | None = None
    fecha_despido: date | None = None
    motivo_despido: str = ""
    fecha_inactividad: date | None = None
    motivo_inactividad: str = ""
    inicio_actividad: date | None = None
    fin_actividad: date | None = None
    dni: str = ""
    extranjero: str = ""
    sexo: str = ""
    cargo: str = ""
    seccion: str = ""
    contrato: str = ""
    unidad_trabajo: str = ""
    perfil_contrato: str = ""

    def estado_en(self, fecha: date) -> str:
        """Determina ACTIVO/INACTIVO para la fecha de proceso.

        Se basa en los campos de alta/despido/inactividad del relatorio.
        Nunca elimina históricos: solo clasifica.
        """
        # Despedido: fecha de despido <= fecha
        if self.fecha_despido and self.fecha_despido <= fecha:
            return "INACTIVO"
        # Inactividad vigente en la fecha
        fi = self.fecha_inactividad or self.inicio_actividad
        ff = self.fin_actividad
        if fi and ff and fi <= fecha <= ff:
            return "INACTIVO"
        if fi and ff is None and fi <= fecha:
            return "INACTIVO"
        # Aún no admitido
        if self.fecha_admision and fecha < self.fecha_admision:
            return "INACTIVO"
        return "ACTIVO"


def leer_relatorio(ruta, cfg, avisos=None) -> list[Empleado]:
    avisos = [] if avisos is None else avisos
    if not os.path.exists(ruta):
        raise IngestaError("Relatorio no encontrado: %s" % ruta)
    rc = cfg.get("relatorio") or {}
    cols_cfg = rc.get("columnas") or {}
    try:
        wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    except Exception as exc:
        raise IngestaError("No se pudo abrir el Relatorio %s (%s)."
                           % (os.path.basename(ruta), str(exc)[:120]))
    try:
        ws = _elegir_hoja(wb, rc.get("hoja") or rc.get("hoja_relatorio"))
        fila_enc = int(rc.get("fila_encabezado", 1))
        mapa = _mapa_encabezados(ws, fila_enc)
        # índice por clave configurada
        idx = {}
        for nombre, columna in cols_cfg.items():
            idx[nombre] = _col(mapa, cols_cfg, columna)

        idx_alcance = {
            "Empleado": idx.get("empleado"),
            "Fotocheck": idx.get("fotocheck"),
            "DNI": idx.get("dni"),
        }
        if idx_alcance["Empleado"] is None:
            raise IngestaError("Relatorio no tiene la columna '%s'." % cols_cfg.get("empleado", "Empleado"))

        out: list[Empleado] = []
        for fila in ws.iter_rows(min_row=fila_enc + 1, values_only=True):
            emp_nombre = _obtener(fila, idx["empleado"])
            if not emp_nombre or not str(emp_nombre).strip():
                continue
            m = Empleado(
                empresa=str(_obtener(fila, idx.get("empresa")) or "").strip(),
                unidad=str(_obtener(fila, idx.get("unidad")) or "").strip(),
                grupo_empresa=str(_obtener(fila, idx.get("grupo_empresa")) or "").strip(),
                codigo_empresa=str(_obtener(fila, idx.get("codigo_empresa")) or "").strip(),
                empresa_terceros=str(_obtener(fila, idx.get("empresa_terceros")) or "").strip(),
                ruc=limpiar_ruc(_obtener(fila, idx.get("ruc"))),
                nombre_comercial=str(_obtener(fila, idx.get("nombre_comercial")) or "").strip(),
                grupo_terceros=str(_obtener(fila, idx.get("grupo_terceros")) or "").strip(),
                empleado=str(emp_nombre).strip(),
                fotocheck=limpiar_codigo(_obtener(fila, idx.get("fotocheck"))),
                fecha_admision=normalizar_fecha(_obtener(fila, idx.get("fecha_admision"))),
                fecha_despido=normalizar_fecha(_obtener(fila, idx.get("fecha_despido"))),
                motivo_despido=str(_obtener(fila, idx.get("motivo_despido")) or "").strip(),
                fecha_inactividad=normalizar_fecha(_obtener(fila, idx.get("fecha_inactividad"))),
                motivo_inactividad=str(_obtener(fila, idx.get("motivo_inactividad")) or "").strip(),
                inicio_actividad=normalizar_fecha(_obtener(fila, idx.get("inicio_actividad"))),
                fin_actividad=normalizar_fecha(_obtener(fila, idx.get("fin_actividad"))),
                dni=normalizar_dni(_obtener(fila, idx.get("dni"))),
                extranjero=str(_obtener(fila, idx.get("extranjero")) or "").strip(),
                sexo=str(_obtener(fila, idx.get("sexo")) or "").strip(),
                cargo=str(_obtener(fila, idx.get("cargo")) or "").strip(),
                seccion=str(_obtener(fila, idx.get("seccion")) or "").strip(),
                contrato=str(_obtener(fila, idx.get("contrato")) or "").strip(),
                unidad_trabajo=str(_obtener(fila, idx.get("unidad_trabajo")) or "").strip(),
                perfil_contrato=str(_obtener(fila, idx.get("perfil_contrato")) or "").strip(),
            )
            out.append(m)
    finally:
        wb.close()
    return out


# ---------------------------------------------------------------------------
# TARIFAS
# ---------------------------------------------------------------------------
@dataclass
class Tarifa:
    id: Any = None
    title: str = ""
    empresa: str = ""
    objeto: str = ""
    ruc: str = ""
    cargo: str = ""
    c25: Decimal = Decimal("0")
    c35: Decimal = Decimal("0")
    c100: Decimal = Decimal("0")
    tipo_item: str = ""

    def columna(self, tipo: str) -> Decimal:
        t = normalizar_texto(tipo)
        if "100" in t:
            return self.c100
        if "35" in t:
            return self.c35
        return self.c25


def leer_tarifas(ruta, cfg, avisos=None) -> list[Tarifa]:
    """Lee el tarifario. Si no se recibe archivo (o no existe), usa el tarifario
    embebido en datos_embebidos.py (inamovible en el motor)."""
    avisos = [] if avisos is None else avisos
    if not ruta or not os.path.exists(ruta):
        return _tarifas_embebidas(cfg)
    tc = cfg.get("tarifas") or {}
    try:
        wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    except Exception as exc:
        raise IngestaError("No se pudo abrir TARIFAS %s (%s)." % (os.path.basename(ruta), str(exc)[:120]))
    try:
        ws = _elegir_hoja(wb, tc.get("hoja"))
        fila_enc = int(tc.get("fila_encabezado", 1))
        mapa = _mapa_encabezados(ws, fila_enc)
        idx = {
            "id": _col(mapa, tc, tc.get("col_id")),
            "title": _col(mapa, tc, tc.get("col_title")),
            "empresa": _col(mapa, tc, tc.get("col_empresa")),
            "objeto": _col(mapa, tc, tc.get("col_objeto")),
            "ruc": _col(mapa, tc, tc.get("col_ruc")),
            "cargo": _col(mapa, tc, tc.get("col_cargo")),
            "c25": _col(mapa, tc, tc.get("col_25")),
            "c35": _col(mapa, tc, tc.get("col_35")),
            "c100": _col(mapa, tc, tc.get("col_100")),
            "tipo_item": _col(mapa, tc, tc.get("col_tipo_item")),
        }
        if idx["cargo"] is None:
            raise IngestaError("TARIFAS no tiene la columna '%s'." % tc.get("col_cargo"))
        out: list[Tarifa] = []
        for fila in ws.iter_rows(min_row=fila_enc + 1, values_only=True):
            cargo = _obtener(fila, idx["cargo"])
            if not cargo or not str(cargo).strip():
                continue
            t = Tarifa(
                id=_obtener(fila, idx["id"]),
                title=str(_obtener(fila, idx["title"]) or "").strip(),
                empresa=str(_obtener(fila, idx["empresa"]) or "").strip(),
                objeto=str(_obtener(fila, idx["objeto"]) or "").strip(),
                ruc=limpiar_ruc(_obtener(fila, idx["ruc"])),
                cargo=str(cargo).strip(),
                c25=Decimal(str(_obtener(fila, idx["c25"]) or 0) or 0),
                c35=Decimal(str(_obtener(fila, idx["c35"]) or 0) or 0),
                c100=Decimal(str(_obtener(fila, idx["c100"]) or 0) or 0),
                tipo_item=str(_obtener(fila, idx["tipo_item"]) or "").strip(),
            )
            out.append(t)
    finally:
        wb.close()
    return out


def _tarifas_embebidas(cfg) -> list[Tarifa]:
    out: list[Tarifa] = []
    for r in _embed.TARIFAS:
        out.append(Tarifa(
            id=r.get("id"),
            title=_t(r.get("title")),
            empresa=_t(r.get("empresa")),
            objeto=_t(r.get("objeto")),
            ruc=limpiar_ruc(r.get("ruc")),
            cargo=_t(r.get("cargo")),
            c25=Decimal(str(r.get("c25") or 0)),
            c35=Decimal(str(r.get("c35") or 0)),
            c100=Decimal(str(r.get("c100") or 0)),
            tipo_item=_t(r.get("tipo_item")),
        ))
    return out


def _t(v):
    return str(v).strip() if v is not None else ""


# ---------------------------------------------------------------------------
# ÁREAS y GERENCIA
# ---------------------------------------------------------------------------
@dataclass
class Area:
    titulo: str = ""
    id: str = ""
    id_gerencia: str = ""
    gerencia: str = ""  # resuelta


@dataclass
class Gerencia:
    titulo: str = ""
    id: str = ""


def leer_areas(ruta, cfg, avisos=None) -> list[Area]:
    avisos = [] if avisos is None else avisos
    if not ruta or not os.path.exists(ruta):
        return _areas_embebidas(cfg)
    ac = cfg.get("areas") or {}
    try:
        wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    except Exception:
        return []
    try:
        ws = _elegir_hoja(wb, ac.get("hoja"))
        mapa = _mapa_encabezados(ws, int(ac.get("fila_encabezado", 1)))
        idx = {
            "titulo": _col(mapa, ac, ac.get("col_titulo")),
            "id": _col(mapa, ac, ac.get("col_id")),
            "id_gerencia": _col(mapa, ac, ac.get("col_id_gerencia")),
        }
        out = []
        for fila in ws.iter_rows(min_row=int(ac.get("fila_encabezado", 1)) + 1, values_only=True):
            titulo = _obtener(fila, idx["titulo"])
            if not titulo:
                continue
            out.append(Area(
                titulo=str(titulo).strip(),
                id=str(_obtener(fila, idx["id"]) or "").strip(),
                id_gerencia=str(_obtener(fila, idx["id_gerencia"]) or "").strip(),
            ))
    finally:
        wb.close()
    return out


def leer_gerencias(ruta, cfg, avisos=None) -> list[Gerencia]:
    avisos = [] if avisos is None else avisos
    if not ruta or not os.path.exists(ruta):
        return _gerencias_embebidas()
    gc = cfg.get("gerencia") or {}
    try:
        wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    except Exception:
        return []
    try:
        ws = _elegir_hoja(wb, gc.get("hoja"))
        mapa = _mapa_encabezados(ws, int(gc.get("fila_encabezado", 1)))
        idx = {
            "titulo": _col(mapa, gc, gc.get("col_titulo")),
            "id": _col(mapa, gc, gc.get("col_id")),
        }
        out = []
        for fila in ws.iter_rows(min_row=int(gc.get("fila_encabezado", 1)) + 1, values_only=True):
            titulo = _obtener(fila, idx["titulo"])
            if not titulo:
                continue
            out.append(Gerencia(titulo=str(titulo).strip(),
                                id=str(_obtener(fila, idx["id"]) or "").strip()))
    finally:
        wb.close()
    return out


def resolver_gerencias(areas: list[Area], gerencias: list[Gerencia]) -> None:
    """Enriquece cada Área con el nombre de su gerencia."""
    gmap = {g.id: g.titulo for g in gerencias}
    for a in areas:
        if a.id_gerencia in gmap:
            a.gerencia = gmap[a.id_gerencia]


def _areas_embebidas(cfg) -> list[Area]:
    out = []
    for r in _embed.AREAS:
        out.append(Area(
            titulo=_t(r.get("titulo")),
            id=_t(r.get("id")),
            id_gerencia=_t(r.get("id_gerencia")),
        ))
    return out


def _gerencias_embebidas() -> list[Gerencia]:
    out = []
    for r in _embed.GERENCIAS:
        out.append(Gerencia(titulo=_t(r.get("titulo")), id=_t(r.get("id"))))
    return out
